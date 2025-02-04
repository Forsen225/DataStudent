
# Импортируем библиотеки, указываем путь к таблице, активируем лист

from math import sqrt # для взятия корня 
from openpyxl import load_workbook, Workbook # как по мне удобней интуитивно чем pandas
import base64

# Создаем функцию для зашифровки имени(для ананимности )
def code_name(name):
    return base64.b64encode(name.encode('utf-8')).decode('utf - 8')


# указываем пусть к файлу
file_path = "C: ...\\...\\..."



#Активируем лист excel
wb = load_workbook(file_path)  
sheet = wb.active



columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'] # колонки с оценками и упражнеиями
results = [] # Список с резуьтатами



# Прохоим по строкам с даннфми начиная с 7 и до полсденей заполненой строки (не учитывем группу и назавниае предмета)
for row in sheet.iter_rows(min_row = 7, max_row=sheet.max_row, values_only=True):
    student_name = row[0]
    if not student_name:
        continue
    
    
    #вызвываем функцию на кодирование пока пробегаемся по строкам     
    coded_name = code_name(student_name)
    
    
    # РАСЧЕТ ПОСТАВЛЕННХ ЗАДАЧ
    grades = [row[sheet[col + '1'].column-1] for col in columns] # оценки из столбцов B–K
    
    grades = [grade if isinstance(grade, (int, float)) else 0 for grade in grades] # Проверям тип данных. Преобразуем "н"(текст) и пропуски в 0

    mid_grade = sum(grades)/len(columns) #Среднее 

    sqr_grafe = (sum([((x - mid_grade)**2) for x in grades]))/len(columns) #СКО

    max_grade = max(grades) #Максимальный балд 

    miss = 0 #Пропуски
    for grade in grades:
        if grade in [0, "н", None]:
            miss += 1
            
            # от себя: Итоговая оценка 
    Total_grade = row[14] # Индекс имеет 14 так как отсчет от 0 (столбецц "O")
    if Total_grade == 5:
        grade_category = "Отл."
    elif Total_grade == 4:
        grade_category = "Хор."
    elif Total_grade == 3:
        grade_category = "Удв."
    else:
        grade_category = "Неуд."
    

    # Сохраняем результаты для текущего студента
    results.append({
        
        "ФИО":coded_name,
        "Макс. балл": max_grade,
        "Средний балл": mid_grade,
        "СКО": sqr_grafe,
        "Пропуски": miss,
        "Итог": grade_category
        })




# Создаем новый файл с нашими обработанными данными    
output_wp = Workbook()
output_sheet = output_wp.active
output_sheet.title = "Результаты"


# Добавляем заголовки в файл
headres = ["ФИО", "Макс. балл", "Средний балл", "СКО", "Пропуски", "Итог"]
output_sheet.append(headres)


# Заполняем файл данными
for result in results:
    output_sheet.append([
        result["ФИО"], 
        result["Макс. балл"], 
        result["Средний балл"], 
        result["СКО"], 
        result["Пропуски"],
        result["Итог"]
        ])




# Сохраняем файл
output_file_path = "NewTabl1.xlsx"
output_wp.save(output_file_path)


# Оповестим через терминал
print(f'Скачен {output_file_path} !')
